from __future__ import annotations

import csv
import hashlib
import json
from collections import Counter
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .config import (
    MAX_DATA_CELLS,
    MAX_DATA_COLUMNS,
    MAX_DATA_ROWS,
    MAX_XLSX_UNCOMPRESSED_BYTES,
)


class DataValidationError(ValueError):
    """Raised when an input file cannot be interpreted safely."""


def _validate_shape(rows: int, columns: int) -> None:
    if rows > MAX_DATA_ROWS:
        raise DataValidationError(f"数据行数超过上限（{MAX_DATA_ROWS:,} 行）")
    if columns > MAX_DATA_COLUMNS:
        raise DataValidationError(f"数据列数超过上限（{MAX_DATA_COLUMNS:,} 列）")
    if rows * columns > MAX_DATA_CELLS:
        raise DataValidationError(f"数据单元格数超过上限（{MAX_DATA_CELLS:,} 个）")


def _validate_headers(values: list[Any] | tuple[Any, ...]) -> None:
    headers = ["" if value is None else str(value).strip() for value in values]
    if any(not header for header in headers):
        raise DataValidationError("数据包含空列名")
    duplicates = sorted(
        header for header, count in Counter(headers).items() if count > 1
    )
    if duplicates:
        raise DataValidationError(f"数据包含重复列名: {', '.join(duplicates)}")


def _validate_xlsx_archive(path: Path) -> None:
    try:
        with ZipFile(path) as archive:
            uncompressed_size = sum(item.file_size for item in archive.infolist())
    except BadZipFile as exc:
        raise DataValidationError("XLSX 文件结构无效") from exc
    if uncompressed_size > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise DataValidationError("XLSX 解压后体积超过 200 MB 限制")


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def excel_sheets(path: Path) -> list[str]:
    if path.suffix.lower() != ".xlsx":
        return []
    _validate_xlsx_archive(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def load_dataframe(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        _validate_xlsx_archive(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheets = list(workbook.sheetnames)
            selected = sheet_name or (sheets[0] if sheets else None)
            if not selected or selected not in sheets:
                raise DataValidationError("XLSX 工作表不存在")
            worksheet = workbook[selected]
            _validate_shape(max(worksheet.max_row - 1, 0), worksheet.max_column)
            header = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
            )
            _validate_headers(header)
        finally:
            workbook.close()
        frame = pd.read_excel(path, sheet_name=selected, engine="openpyxl")
    elif suffix == ".csv":
        errors: list[str] = []
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                with path.open("r", encoding=encoding, newline="") as stream:
                    header = next(csv.reader(stream), None)
                if header is None:
                    raise DataValidationError("数据文件没有表头")
                _validate_headers(header)
                chunks: list[pd.DataFrame] = []
                row_count = 0
                reader = pd.read_csv(path, encoding=encoding, chunksize=25_000)
                for chunk in reader:
                    row_count += len(chunk)
                    _validate_shape(row_count, chunk.shape[1])
                    chunks.append(chunk)
                frame = (
                    pd.concat(chunks, ignore_index=True)
                    if chunks
                    else pd.read_csv(path, encoding=encoding, nrows=0)
                )
                break
            except UnicodeDecodeError as exc:
                errors.append(str(exc))
        else:
            raise DataValidationError("CSV 编码无法识别，请另存为 UTF-8 或 GB18030") from None
    else:
        raise DataValidationError("仅支持 .csv 与 .xlsx 文件")

    frame.columns = [str(column).strip() for column in frame.columns]
    if any(not column for column in frame.columns):
        raise DataValidationError("数据包含空列名")
    duplicates = frame.columns[frame.columns.duplicated()].tolist()
    if duplicates:
        raise DataValidationError(f"数据包含重复列名: {', '.join(duplicates)}")
    if frame.empty:
        raise DataValidationError("数据文件没有可分析的记录")
    _validate_shape(len(frame), frame.shape[1])
    return frame


def json_value(value: Any) -> Any:
    if value is None or value is pd.NA:
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        return None if not np.isfinite(value) else float(value)
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, (pd.Timestamp,)):
        return value.isoformat()
    return value


def records_for_json(frame: pd.DataFrame, limit: int | None = None) -> list[dict[str, Any]]:
    source = frame.head(limit) if limit is not None else frame
    return [
        {str(key): json_value(value) for key, value in row.items()}
        for row in source.to_dict(orient="records")
    ]


def inspect_dataframe(frame: pd.DataFrame) -> dict[str, Any]:
    missing = frame.isna().sum()
    numeric = frame.apply(lambda series: pd.to_numeric(series, errors="coerce").notna().mean())
    return {
        "rows": int(len(frame)),
        "columns_count": int(frame.shape[1]),
        "columns": [str(column) for column in frame.columns],
        "preview": records_for_json(frame, limit=8),
        "quality": {
            "duplicate_rows": int(frame.duplicated().sum()),
            "empty_columns": [str(column) for column in frame.columns if frame[column].isna().all()],
            "constant_columns": [
                str(column) for column in frame.columns if frame[column].nunique(dropna=True) <= 1
            ],
            "missing_cells": int(missing.sum()),
            "missing_by_column": {
                str(column): int(count) for column, count in missing.items() if count > 0
            },
            "numeric_share": {
                str(column): round(float(share), 4) for column, share in numeric.items()
            },
        },
    }


def write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, allow_nan=False, default=json_value),
        encoding="utf-8",
    )
