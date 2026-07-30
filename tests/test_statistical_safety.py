from __future__ import annotations

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook
from semopy import Model

from app.reporting import export_excel
from app.schemas import AnalysisRequest, RoleConfig, ScaleConfig
from app.sem_analysis import _latent_diagnostics, _srmr, run_cfa, run_harman
from app.stats_utils import bootstrap_statistics, fit_ols, least_squares_coefficient


def test_ols_confidence_interval_uses_requested_alpha_and_vif_has_intercept():
    rng = np.random.default_rng(20260730)
    rows = 400
    x1 = rng.normal(10, 1, rows)
    x2 = rng.normal(-6, 1, rows)
    frame = pd.DataFrame(
        {
            "x1": x1,
            "x2": x2,
            "y": 2 + 0.5 * x1 - 0.2 * x2 + rng.normal(0, 1, rows),
        }
    )

    _, summary_95 = fit_ols(frame, "y", ["x1", "x2"], alpha=0.05)
    _, summary_80 = fit_ols(frame, "y", ["x1", "x2"], alpha=0.20)
    coefficient_95 = next(row for row in summary_95["coefficients"] if row["term"] == "x1")
    coefficient_80 = next(row for row in summary_80["coefficients"] if row["term"] == "x1")

    assert coefficient_80["ci_high"] - coefficient_80["ci_low"] < (
        coefficient_95["ci_high"] - coefficient_95["ci_low"]
    )
    assert summary_80["confidence_level"] == pytest.approx(0.8)
    assert max(row["vif"] for row in summary_95["vif"]) < 1.1


def test_cfa_rejects_nonpositive_degrees_of_freedom():
    rng = np.random.default_rng(11)
    item_data = pd.DataFrame(rng.normal(size=(120, 3)), columns=["q1", "q2", "q3"])
    scales = [ScaleConfig(name="单构念", items=["q1", "q2", "q3"])]

    with pytest.raises(ValueError, match="自由度必须大于 0"):
        run_cfa(item_data, scales)


def test_harman_uses_complete_cases():
    rng = np.random.default_rng(19)
    item_data = pd.DataFrame(rng.normal(size=(12, 3)), columns=["q1", "q2", "q3"])
    item_data.loc[0, "q2"] = np.nan

    result = run_harman(item_data, 40)

    assert result["n"] == 11
    assert result["missing"] == "listwise"


def test_excel_export_does_not_interpret_user_strings_as_formulas(tmp_path):
    output_path = tmp_path / "safe.xlsx"
    export_excel(
        {
            "run_id": "=1+1",
            "created_at": "2026-07-30",
            "data_sha256": "abc",
            "data_quality": {"input_rows": 1},
            "errors": [],
        },
        output_path,
    )

    workbook = load_workbook(output_path, data_only=False)
    cell = workbook["运行概览"]["B2"]
    assert cell.value == "=1+1"
    assert cell.data_type == "s"


def test_exact_collinearity_is_rejected_in_models_and_bootstrap_helpers():
    frame = pd.DataFrame(
        {
            "x1": np.arange(30, dtype=float),
            "x2": np.arange(30, dtype=float) * 2,
            "y": np.arange(30, dtype=float) + 1,
        }
    )

    with pytest.raises(ValueError, match="完全共线性"):
        fit_ols(frame, "y", ["x1", "x2"])
    with pytest.raises(ValueError, match="完全共线性"):
        least_squares_coefficient(frame, "y", ["x1", "x2"], "x1")


def test_ols_is_stable_when_predictor_has_a_large_location_shift():
    rng = np.random.default_rng(808)
    centered_x = rng.normal(0, 1, 500)
    x = 1e14 + centered_x
    frame = pd.DataFrame(
        {"x": x, "y": 4 + 0.6 * centered_x + rng.normal(0, 0.2, len(x))}
    )

    _, summary = fit_ols(frame, "y", ["x"])
    slope = next(row for row in summary["coefficients"] if row["term"] == "x")

    assert slope["b"] == pytest.approx(0.6, abs=0.03)


def test_bca_requires_every_delete_one_jackknife_fit_to_succeed():
    row_count = 30

    def evaluator(index):
        if len(index) < row_count:
            raise ValueError("jackknife failure")
        return {"effect": 1.0}

    with pytest.raises(ValueError, match="Jackknife 未全部成功"):
        bootstrap_statistics(
            row_count,
            evaluator,
            samples=200,
            seed=1,
            alpha=0.05,
            interval="bca",
        )


def test_bca_rejects_pathological_jackknife_size_before_sampling():
    with pytest.raises(ValueError, match="BCa 区间最多支持"):
        bootstrap_statistics(
            2_501,
            lambda index: {"effect": float(len(index))},
            samples=200,
            seed=1,
            alpha=0.05,
            interval="bca",
        )


def test_percentile_bootstrap_rejects_excessive_compute_budget():
    with pytest.raises(ValueError, match="Bootstrap 计算量超过上限"):
        bootstrap_statistics(
            50_001,
            lambda index: {"effect": float(len(index))},
            samples=200,
            seed=1,
            alpha=0.05,
            interval="percentile",
        )


def test_schema_rejects_ambiguous_roles_and_cross_scale_items():
    with pytest.raises(ValueError, match="必须使用不同变量"):
        RoleConfig(x="x", y="x")

    with pytest.raises(ValueError, match="同时属于多个量表"):
        AnalysisRequest.model_validate(
            {
                "dataset_id": "0" * 32,
                "scales": [
                    {"name": "A", "items": ["q1", "q2"]},
                    {"name": "B", "items": ["q2", "q3"]},
                ],
                "roles": {"x": "x", "y": "y", "mediator": "m", "moderator": "w"},
            }
        )


def test_srmr_uses_the_same_n_denominator_covariance_as_mlw():
    rng = np.random.default_rng(91)
    rows = 40
    f1 = rng.normal(size=rows)
    f2 = rng.normal(size=rows)
    data = pd.DataFrame(
        {
            "a1": f1 + rng.normal(0, 0.8, rows),
            "a2": 0.8 * f1 + rng.normal(0, 0.8, rows),
            "a3": 0.7 * f1 + rng.normal(0, 0.8, rows),
            "b1": f2 + rng.normal(0, 0.8, rows),
            "b2": 0.8 * f2 + rng.normal(0, 0.8, rows),
            "b3": 0.7 * f2 + rng.normal(0, 0.8, rows),
        }
    )
    model = Model("f1 =~ a1 + a2 + a3\nf2 =~ b1 + b2 + b3")
    model.fit(data, obj="MLW")
    sample = np.asarray(model.mx_cov, dtype=float)
    implied = np.asarray(model.calc_sigma()[0], dtype=float)
    standardizer = np.sqrt(np.outer(np.diag(sample), np.diag(sample)))
    residual = (sample - implied) / standardizer
    expected = np.sqrt(np.mean(np.square(residual[np.tril_indices_from(residual)])))

    assert _srmr(model, data) == pytest.approx(expected)


def test_improper_latent_correlation_is_detected():
    class FakeModel:
        vars = {"latent": {"f1", "f2"}}
        mx_psi = np.asarray([[1.0, 1.05], [1.05, 1.0]])

    estimates = pd.DataFrame(
        [
            {"lval": "f1", "op": "~~", "rval": "f2", "Est. Std": 1.05},
            {"lval": "f1", "op": "~~", "rval": "f1", "Est. Std": 1.0},
            {"lval": "f2", "op": "~~", "rval": "f2", "Est. Std": 1.0},
        ]
    )

    diagnostics = _latent_diagnostics(FakeModel(), estimates)

    assert diagnostics["improper_latent_correlation"] is True
    assert diagnostics["latent_covariance_positive_definite"] is False


def test_ulmc_only_export_still_contains_sem_fit_sheet(tmp_path):
    output_path = tmp_path / "ulmc.xlsx"
    fit = {"n": 100, "df": 20, "cfi": 0.95, "tli": 0.94, "rmsea": 0.05}
    export_excel(
        {
            "run_id": "run-000000000000",
            "created_at": "2026-07-30",
            "data_sha256": "abc",
            "data_quality": {"input_rows": 100},
            "errors": [],
            "ulmc": {
                "status": "ok",
                "trait_only_fit": fit,
                "trait_method_fit": fit,
                "comparison": {},
                "method_loadings": [],
            },
        },
        output_path,
    )

    workbook = load_workbook(output_path, data_only=False)
    assert "SEM拟合" in workbook.sheetnames
    assert workbook["SEM拟合"].max_row == 3
